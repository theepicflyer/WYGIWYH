import csv
import hashlib
import logging
import os
import re
import zipfile
from django.db import transaction
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import Dict, Any, Literal, Union

import openpyxl
import xlrd
import yaml
from cachalot.api import cachalot_disabled
from django.conf import settings
from django.utils import timezone
from openpyxl.utils.exceptions import InvalidFileException

from apps.accounts.models import Account, AccountGroup
from apps.currencies.models import Currency
from apps.import_app.models import ImportRun, ImportProfile
from apps.import_app.schemas import version_1
from apps.transactions.models import (
    Transaction,
    TransactionCategory,
    TransactionTag,
    TransactionEntity,
)
from apps.rules.signals import transaction_created
from apps.import_app.schemas.v1 import (
    TransactionCategoryMapping,
    TransactionAccountMapping,
    TransactionTagsMapping,
    TransactionEntitiesMapping,
)

logger = logging.getLogger(__name__)


class ImportService:
    TEMP_DIR = os.getenv("IMPORT_TEMP_DIR", str(settings.BASE_DIR / "temp"))

    def __init__(self, import_run: ImportRun):
        self.import_run: ImportRun = import_run
        self.profile: ImportProfile = import_run.profile
        self.config: version_1.ImportProfileSchema = self._load_config()
        self.settings: version_1.CSVImportSettings | version_1.ExcelImportSettings = (
            self.config.settings
        )
        self.deduplication: list[version_1.CompareDeduplicationRule] = (
            self.config.deduplication
        )
        self.mapping: Dict[str, version_1.ColumnMapping] = self.config.mapping

        # Ensure temp directory exists
        os.makedirs(self.TEMP_DIR, exist_ok=True)

    def _load_config(self) -> version_1.ImportProfileSchema:
        yaml_data = yaml.safe_load(self.profile.yaml_config)
        try:
            config = version_1.ImportProfileSchema(**yaml_data)
        except Exception as e:
            self._log("error", f"Fatal error processing YAML config: {str(e)}")
            self._update_status("FAILED")
            raise e
        else:
            return config

    def _log(self, level: str, message: str, **kwargs) -> None:
        """Add a log entry to the import run logs"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Format additional context if present
        context = ""
        if kwargs:
            context = " - " + ", ".join(f"{k}={v}" for k, v in kwargs.items())

        log_line = f"[{timestamp}] {level.upper()}: {message}{context}\n"

        # Append to existing logs
        self.import_run.logs += log_line
        self.import_run.save(update_fields=["logs"])

        if level == "info":
            logger.info(log_line)
        elif level == "warning":
            logger.warning(log_line)
        elif level == "error":
            logger.error(log_line, exc_info=True)

    def _update_totals(
        self,
        field: Literal["total", "processed", "successful", "skipped", "failed"],
        value: int,
    ) -> None:
        if field == "total":
            self.import_run.total_rows = value
            self.import_run.save(update_fields=["total_rows"])
        elif field == "processed":
            self.import_run.processed_rows = value
            self.import_run.save(update_fields=["processed_rows"])
        elif field == "successful":
            self.import_run.successful_rows = value
            self.import_run.save(update_fields=["successful_rows"])
        elif field == "skipped":
            self.import_run.skipped_rows = value
            self.import_run.save(update_fields=["skipped_rows"])
        elif field == "failed":
            self.import_run.failed_rows = value
            self.import_run.save(update_fields=["failed_rows"])

    def _increment_totals(
        self,
        field: Literal["total", "processed", "successful", "skipped", "failed"],
        value: int,
    ) -> None:
        if field == "total":
            self.import_run.total_rows = self.import_run.total_rows + value
            self.import_run.save(update_fields=["total_rows"])
        elif field == "processed":
            self.import_run.processed_rows = self.import_run.processed_rows + value
            self.import_run.save(update_fields=["processed_rows"])
        elif field == "successful":
            self.import_run.successful_rows = self.import_run.successful_rows + value
            self.import_run.save(update_fields=["successful_rows"])
        elif field == "skipped":
            self.import_run.skipped_rows = self.import_run.skipped_rows + value
            self.import_run.save(update_fields=["skipped_rows"])
        elif field == "failed":
            self.import_run.failed_rows = self.import_run.failed_rows + value
            self.import_run.save(update_fields=["failed_rows"])

    def _update_status(
        self, new_status: Literal["PROCESSING", "FAILED", "FINISHED"]
    ) -> None:
        if new_status == "PROCESSING":
            self.import_run.status = ImportRun.Status.PROCESSING
        elif new_status == "FAILED":
            self.import_run.status = ImportRun.Status.FAILED
        elif new_status == "FINISHED":
            self.import_run.status = ImportRun.Status.FINISHED

        self.import_run.save(update_fields=["status"])

    def _transform_value(
        self,
        value: str,
        mapping: version_1.ColumnMapping,
        row: Dict[str, str] = None,
        mapped_data: Dict[str, Any] = None,
    ) -> Any:
        transformed = value

        for transform in mapping.transformations:
            if transform.type == "hash":
                # Collect all values to be hashed
                values_to_hash = []
                for field in transform.fields:
                    if field in row:
                        values_to_hash.append(str(row[field]))
                    elif (
                        field.startswith("__")
                        and mapped_data
                        and field[2:] in mapped_data
                    ):
                        values_to_hash.append(str(mapped_data[field[2:]]))
                if values_to_hash:
                    concatenated = "|".join(values_to_hash)
                    transformed = hashlib.sha256(concatenated.encode()).hexdigest()

            elif transform.type == "replace":
                if transform.exclusive:
                    transformed = value.replace(
                        transform.pattern, transform.replacement
                    )
                else:
                    transformed = transformed.replace(
                        transform.pattern, transform.replacement
                    )

            elif transform.type == "regex":
                if transform.exclusive:
                    transformed = re.sub(
                        transform.pattern, transform.replacement, value
                    )
                else:
                    transformed = re.sub(
                        transform.pattern, transform.replacement, transformed
                    )

            elif transform.type == "date_format":
                transformed = datetime.strptime(
                    transformed, transform.original_format
                ).strftime(transform.new_format)

            elif transform.type == "merge":
                values_to_merge = []
                for field in transform.fields:
                    if field in row:
                        values_to_merge.append(str(row[field]))
                    elif (
                        field.startswith("__")
                        and mapped_data
                        and field[2:] in mapped_data
                    ):
                        values_to_merge.append(str(mapped_data[field[2:]]))
                transformed = transform.separator.join(values_to_merge)

            elif transform.type == "split":
                parts = transformed.split(transform.separator)
                if transform.index is not None:
                    transformed = parts[transform.index] if parts else ""
                else:
                    transformed = parts

            elif transform.type in ["add", "subtract"]:
                try:
                    source_value = Decimal(transformed)

                    # First check row data, then mapped data if not found
                    field_value = row.get(transform.field)
                    if field_value is None and transform.field.startswith("__"):
                        field_value = mapped_data.get(transform.field[2:])

                    if field_value is None:
                        raise KeyError(
                            f"Field '{transform.field}' not found in row or mapped data"
                        )

                    field_value = self._prepare_numeric_value(
                        str(field_value),
                        transform.thousand_separator,
                        transform.decimal_separator,
                    )

                    if transform.absolute_values:
                        source_value = abs(source_value)
                        field_value = abs(field_value)

                    if transform.type == "add":
                        transformed = str(source_value + field_value)
                    else:  # subtract
                        transformed = str(source_value - field_value)
                except (InvalidOperation, KeyError, AttributeError) as e:
                    logger.warning(
                        f"Error in {transform.type} transformation: {e}. Values: {transformed}, {transform.field}"
                    )
        return transformed

    def _create_transaction(self, data: Dict[str, Any]) -> Transaction:
        tags = []
        entities = []
        # Handle related objects first
        if "category" in data:
            if "category" in data:
                category_name = data.pop("category")
                category_mapping = next(
                    (
                        m
                        for m in self.mapping.values()
                        if isinstance(m, TransactionCategoryMapping)
                        and m.target == "category"
                    ),
                    None,
                )

                try:
                    if category_mapping:
                        if category_mapping.type == "id":
                            category = TransactionCategory.objects.get(id=category_name)
                        else:  # name
                            if getattr(category_mapping, "create", False):
                                try:
                                    category = TransactionCategory.objects.get(
                                        name=category_name
                                    )
                                except TransactionCategory.DoesNotExist:
                                    category = TransactionCategory(name=category_name)
                                    category.save()
                            else:
                                category = TransactionCategory.objects.filter(
                                    name=category_name
                                ).first()
                        if category:
                            data["category"] = category
                            self.import_run.categories.add(category)
                except (TransactionCategory.DoesNotExist, ValueError):
                    # Ignore if category doesn't exist and create is False or not set
                    data["category"] = None

        if "account" in data:
            account_id = data.pop("account")
            account_mapping = next(
                (
                    m
                    for m in self.mapping.values()
                    if isinstance(m, TransactionAccountMapping)
                    and m.target == "account"
                ),
                None,
            )

            try:
                if account_mapping and account_mapping.type == "id":
                    account = Account.objects.filter(id=account_id).first()
                else:  # name
                    account = Account.objects.filter(name=account_id).first()

                if account:
                    data["account"] = account
            except ValueError:
                # Ignore if account doesn't exist
                pass

        if "tags" in data:
            tag_names = data.pop("tags")
            tags_mapping = next(
                (
                    m
                    for m in self.mapping.values()
                    if isinstance(m, TransactionTagsMapping) and m.target == "tags"
                ),
                None,
            )

            for tag_name in tag_names:
                try:
                    if tags_mapping:
                        if tags_mapping.type == "id":
                            tag = TransactionTag.objects.filter(id=tag_name).first()
                        else:  # name
                            if getattr(tags_mapping, "create", False):
                                try:
                                    tag = TransactionTag.objects.get(
                                        name=tag_name.strip()
                                    )
                                except TransactionTag.DoesNotExist:
                                    tag = TransactionTag(name=tag_name.strip())
                                    tag.save()
                            else:
                                tag = TransactionTag.objects.filter(
                                    name=tag_name.strip()
                                ).first()

                        if tag:
                            tags.append(tag)
                            self.import_run.tags.add(tag)
                except ValueError:
                    # Ignore if tag doesn't exist and create is False or not set
                    continue

        if "entities" in data:
            entity_names = data.pop("entities")
            entities_mapping = next(
                (
                    m
                    for m in self.mapping.values()
                    if isinstance(m, TransactionEntitiesMapping)
                    and m.target == "entities"
                ),
                None,
            )

            for entity_name in entity_names:
                try:
                    if entities_mapping:
                        if entities_mapping.type == "id":
                            entity = TransactionTag.objects.filter(
                                id=entity_name
                            ).first()
                        else:  # name
                            if getattr(entities_mapping, "create", False):
                                try:
                                    entity = TransactionEntity.objects.get(
                                        name=entity_name.strip()
                                    )
                                except TransactionEntity.DoesNotExist:
                                    entity = TransactionEntity(name=entity_name.strip())
                                    entity.save()
                            else:
                                entity = TransactionEntity.objects.filter(
                                    name=entity_name.strip()
                                ).first()

                        if entity:
                            entities.append(entity)
                            self.import_run.entities.add(entity)
                except ValueError:
                    # Ignore if entity doesn't exist and create is False or not set
                    continue

        # Create the transaction
        new_transaction = Transaction.objects.create(**data)
        self.import_run.transactions.add(new_transaction)

        # Add many-to-many relationships
        if tags:
            new_transaction.tags.set(tags)
        if entities:
            new_transaction.entities.set(entities)

        if self.settings.trigger_transaction_rules:
            transaction_created.send(sender=new_transaction)

        return new_transaction

    def _create_account(self, data: Dict[str, Any]) -> Account:
        if "group" in data:
            group_name = data.pop("group")
            try:
                group = AccountGroup.objects.get(name=group_name)
            except AccountGroup.DoesNotExist:
                group = AccountGroup(name=group_name)
                group.save()
            data["group"] = group

        # Handle currency references
        if "currency" in data:
            currency = Currency.objects.get(code=data["currency"])
            data["currency"] = currency
            self.import_run.currencies.add(currency)

        if "exchange_currency" in data:
            exchange_currency = Currency.objects.get(code=data["exchange_currency"])
            data["exchange_currency"] = exchange_currency
            self.import_run.currencies.add(exchange_currency)

        return Account.objects.create(**data)

    def _create_currency(self, data: Dict[str, Any]) -> Currency:
        # Handle exchange currency reference
        if "exchange_currency" in data:
            exchange_currency = Currency.objects.get(code=data["exchange_currency"])
            data["exchange_currency"] = exchange_currency
            self.import_run.currencies.add(exchange_currency)

        currency = Currency.objects.create(**data)
        self.import_run.currencies.add(currency)
        return currency

    def _create_category(self, data: Dict[str, Any]) -> TransactionCategory:
        category = TransactionCategory.objects.create(**data)
        self.import_run.categories.add(category)
        return category

    def _create_tag(self, data: Dict[str, Any]) -> TransactionTag:
        tag = TransactionTag.objects.create(**data)
        self.import_run.tags.add(tag)
        return tag

    def _create_entity(self, data: Dict[str, Any]) -> TransactionEntity:
        entity = TransactionEntity.objects.create(**data)
        self.import_run.entities.add(entity)
        return entity

    def _check_duplicate_transaction(self, transaction_data: Dict[str, Any]) -> bool:
        for rule in self.deduplication:
            if rule.type == "compare":
                query = Transaction.all_objects.all().values("id")

                # Build query conditions for each field in the rule
                for field in rule.fields:
                    if field in transaction_data:
                        value = transaction_data[field]
                        # Use __iexact only for string fields; non-string types
                        # (date, Decimal, bool, int, etc.) don't support UPPER()
                        if rule.match_type == "strict" or not isinstance(value, str):
                            query = query.filter(**{field: value})
                        else:  # lax matching for strings only
                            query = query.filter(**{f"{field}__iexact": value})

                # If we found any matching transaction, it's a duplicate
                if query.exists():
                    return True

        return False

    def _coerce_type(
        self, value: str, mapping: version_1.ColumnMapping
    ) -> Union[str, int, bool, Decimal, datetime, list, None]:
        coerce_to = mapping.coerce_to

        # Handle detection methods that don't require a source value
        if coerce_to == "transaction_type" and isinstance(
            mapping, version_1.TransactionTypeMapping
        ):
            if mapping.detection_method == "always_income":
                return Transaction.Type.INCOME
            elif mapping.detection_method == "always_expense":
                return Transaction.Type.EXPENSE
        elif coerce_to == "is_paid" and isinstance(
            mapping, version_1.TransactionIsPaidMapping
        ):
            if mapping.detection_method == "always_paid":
                return True
            elif mapping.detection_method == "always_unpaid":
                return False

        if not value:
            return None

        return self._coerce_single_type(value, coerce_to, mapping)

    @staticmethod
    def _coerce_single_type(
        value: str, coerce_to: str, mapping: version_1.ColumnMapping
    ) -> Union[str, int, bool, Decimal, datetime.date, list]:
        if coerce_to == "str":
            return str(value)
        elif coerce_to == "int":
            return int(value)
        elif coerce_to == "str|int":
            if hasattr(mapping, "type") and mapping.type == "id":
                return int(value)
            elif hasattr(mapping, "type") and mapping.type in ["name", "code"]:
                return str(value)
            else:
                return str(value)
        elif coerce_to == "bool":
            return value.lower() in ["true", "1", "yes", "y", "on"]
        elif coerce_to == "positive_decimal":
            return abs(Decimal(value))
        elif coerce_to == "date":
            if isinstance(
                mapping,
                (
                    version_1.TransactionDateMapping,
                    version_1.TransactionReferenceDateMapping,
                ),
            ):
                if isinstance(value, datetime):
                    return value.date()
                elif isinstance(value, date):
                    return value

                formats = (
                    mapping.format
                    if isinstance(mapping.format, list)
                    else [mapping.format]
                )
                for fmt in formats:
                    try:
                        return datetime.strptime(value, fmt).date()
                    except ValueError:
                        continue
                raise ValueError(
                    f"Could not parse date '{value}' with any of the provided formats"
                )
            else:
                raise ValueError(
                    "Date coercion is only supported for TransactionDateMapping and TransactionReferenceDateMapping"
                )
        elif coerce_to == "list":
            return (
                value
                if isinstance(value, list)
                else [item.strip() for item in value.split(",") if item.strip()]
            )
        elif coerce_to == "transaction_type":
            if isinstance(mapping, version_1.TransactionTypeMapping):
                if mapping.detection_method == "sign":
                    return (
                        Transaction.Type.EXPENSE
                        if value.startswith("-")
                        else Transaction.Type.INCOME
                    )
                elif mapping.detection_method == "always_income":
                    return Transaction.Type.INCOME
                elif mapping.detection_method == "always_expense":
                    return Transaction.Type.EXPENSE
            raise ValueError("Invalid transaction type detection method")
        elif coerce_to == "is_paid":
            if isinstance(mapping, version_1.TransactionIsPaidMapping):
                if mapping.detection_method == "boolean":
                    return value.lower() in ["true", "1", "yes", "y", "on"]
                elif mapping.detection_method == "always_paid":
                    return True
                elif mapping.detection_method == "always_unpaid":
                    return False
            raise ValueError("Invalid is_paid detection method")
        else:
            raise ValueError(f"Unsupported coercion type: {coerce_to}")

    def _map_row(self, row: Dict[str, str]) -> Dict[str, Any]:
        mapped_data = {}
        for field, mapping in self.mapping.items():
            value = None
            if isinstance(mapping.source, str):
                if mapping.source in row:
                    value = row[mapping.source]
                elif (
                    mapping.source.startswith("__")
                    and mapping.source[2:] in mapped_data
                ):
                    value = mapped_data[mapping.source[2:]]
            elif isinstance(mapping.source, list):
                for source in mapping.source:
                    if source in row:
                        value = row[source]
                        break
                    elif source.startswith("__") and source[2:] in mapped_data:
                        value = mapped_data[source[2:]]
                        break

            if value is None:
                value = mapping.default

            if mapping.transformations:
                value = self._transform_value(value, mapping, row, mapped_data)

            value = self._coerce_type(value, mapping)

            if mapping.required and value is None:
                raise ValueError(f"Required field {field} is missing")

            if value is not None:
                target = mapping.target
                if self.settings.importing == "transactions":
                    mapped_data[target] = value
                else:
                    field_name = target.split("_", 1)[1]
                    mapped_data[field_name] = value

        return mapped_data

    @staticmethod
    def _prepare_numeric_value(
        value: str, thousand_separator: str, decimal_separator: str
    ) -> Decimal:
        # Remove thousand separators
        if thousand_separator:
            value = value.replace(thousand_separator, "")

        # Replace decimal separator with dot
        if decimal_separator != ".":
            value = value.replace(decimal_separator, ".")

        return Decimal(value)

    def _process_row(self, row: Dict[str, str], row_number: int) -> None:
        try:
            mapped_data = self._map_row(row)

            if mapped_data:
                # Handle different import types
                if self.settings.importing == "transactions":
                    if self.deduplication and self._check_duplicate_transaction(
                        mapped_data
                    ):
                        self._increment_totals("skipped", 1)
                        self._log("info", f"Skipped duplicate row {row_number}")
                        return
                    self._create_transaction(mapped_data)
                elif self.settings.importing == "accounts":
                    self._create_account(mapped_data)
                elif self.settings.importing == "currencies":
                    self._create_currency(mapped_data)
                elif self.settings.importing == "categories":
                    self._create_category(mapped_data)
                elif self.settings.importing == "tags":
                    self._create_tag(mapped_data)
                elif self.settings.importing == "entities":
                    self._create_entity(mapped_data)

                self._increment_totals("successful", value=1)
                self._log("info", f"Successfully processed row {row_number}")

            self._increment_totals("processed", value=1)

        except Exception as e:
            if not self.settings.skip_errors:
                self._log("error", f"Fatal error processing row {row_number}: {str(e)}")
                self._update_status("FAILED")
                raise
            else:
                self._log("warning", f"Error processing row {row_number}: {str(e)}")
                self._increment_totals("failed", value=1)

            logger.error(f"Fatal error processing row {row_number}", exc_info=e)

    def _process_csv(self, file_path):
        # First pass: count rows
        with open(file_path, "r", encoding=self.settings.encoding) as csv_file:
            # Skip specified number of rows
            for _ in range(self.settings.skip_lines):
                next(csv_file)

            reader = csv.DictReader(csv_file, delimiter=self.settings.delimiter)
            self._update_totals("total", value=sum(1 for _ in reader))

        with open(file_path, "r", encoding=self.settings.encoding) as csv_file:
            # Skip specified number of rows
            for _ in range(self.settings.skip_lines):
                next(csv_file)
            if self.settings.skip_lines:
                self._log("info", f"Skipped {self.settings.skip_lines} initial lines")

            reader = csv.DictReader(csv_file, delimiter=self.settings.delimiter)

            self._log("info", f"Starting import with {self.import_run.total_rows} rows")

            for row_number, row in enumerate(reader, start=1):
                self._process_row(row, row_number)

    def _process_excel(self, file_path):
        try:
            if self.settings.file_type == "xlsx":
                workbook = openpyxl.load_workbook(
                    file_path, read_only=True, data_only=True
                )
                sheets_to_process = (
                    workbook.sheetnames
                    if self.settings.sheets == "*"
                    else (
                        self.settings.sheets
                        if isinstance(self.settings.sheets, list)
                        else [self.settings.sheets]
                    )
                )

                # Calculate total rows
                total_rows = sum(
                    max(0, workbook[sheet_name].max_row - self.settings.start_row)
                    for sheet_name in sheets_to_process
                    if sheet_name in workbook.sheetnames
                )
                self._update_totals("total", value=total_rows)

                # Process sheets
                for sheet_name in sheets_to_process:
                    if sheet_name not in workbook.sheetnames:
                        self._log(
                            "warning",
                            f"Sheet '{sheet_name}' not found in the Excel file. Skipping.",
                        )
                        continue

                    sheet = workbook[sheet_name]
                    self._log("info", f"Processing sheet: {sheet_name}")
                    headers = [
                        str(cell.value or "") for cell in sheet[self.settings.start_row]
                    ]

                    for row_number, row in enumerate(
                        sheet.iter_rows(
                            min_row=self.settings.start_row + 1, values_only=True
                        ),
                        start=1,
                    ):
                        try:
                            row_data = {
                                key: str(value) if value is not None else None
                                for key, value in zip(headers, row)
                            }
                            self._process_row(row_data, row_number)
                        except Exception as e:
                            if self.settings.skip_errors:
                                self._log(
                                    "warning",
                                    f"Error processing row {row_number} in sheet '{sheet_name}': {str(e)}",
                                )
                                self._increment_totals("failed", value=1)
                            else:
                                raise

                workbook.close()

            else:  # xls
                workbook = xlrd.open_workbook(file_path)
                sheets_to_process = (
                    workbook.sheet_names()
                    if self.settings.sheets == "*"
                    else (
                        self.settings.sheets
                        if isinstance(self.settings.sheets, list)
                        else [self.settings.sheets]
                    )
                )
                # Calculate total rows
                total_rows = sum(
                    max(
                        0,
                        workbook.sheet_by_name(sheet_name).nrows
                        - self.settings.start_row,
                    )
                    for sheet_name in sheets_to_process
                    if sheet_name in workbook.sheet_names()
                )
                self._update_totals("total", value=total_rows)
                # Process sheets
                for sheet_name in sheets_to_process:
                    if sheet_name not in workbook.sheet_names():
                        self._log(
                            "warning",
                            f"Sheet '{sheet_name}' not found in the Excel file. Skipping.",
                        )
                        continue
                    sheet = workbook.sheet_by_name(sheet_name)
                    self._log("info", f"Processing sheet: {sheet_name}")
                    headers = [
                        str(sheet.cell_value(self.settings.start_row - 1, col) or "")
                        for col in range(sheet.ncols)
                    ]
                    for row_number in range(self.settings.start_row, sheet.nrows):
                        try:
                            row_data = {}
                            for col, key in enumerate(headers):
                                cell_type = sheet.cell_type(row_number, col)
                                cell_value = sheet.cell_value(row_number, col)

                                if cell_type == xlrd.XL_CELL_DATE:
                                    # Convert Excel date to Python datetime
                                    try:
                                        python_date = datetime(
                                            *xlrd.xldate_as_tuple(
                                                cell_value, workbook.datemode
                                            )
                                        )
                                        row_data[key] = python_date
                                    except Exception:
                                        # If date conversion fails, use the original value
                                        row_data[key] = (
                                            str(cell_value)
                                            if cell_value is not None
                                            else None
                                        )
                                elif cell_value is None:
                                    row_data[key] = None
                                else:
                                    row_data[key] = str(cell_value)

                            self._process_row(
                                row_data, row_number - self.settings.start_row + 1
                            )
                        except Exception as e:
                            if self.settings.skip_errors:
                                self._log(
                                    "warning",
                                    f"Error processing row {row_number} in sheet '{sheet_name}': {str(e)}",
                                )
                                self._increment_totals("failed", value=1)
                            else:
                                raise

        except (InvalidFileException, xlrd.XLRDError) as e:
            raise ValueError(
                f"Invalid {self.settings.file_type.upper()} file format: {str(e)}"
            )

    def _parse_and_import_qif(self, content_lines: list[str], filename: str) -> None:
        # Infer account from filename (remove extension)
        account_name = os.path.splitext(os.path.basename(filename))[0]

        current_transaction = {}
        raw_lines_buffer = []

        account = Account.objects.filter(name=account_name).first()
        if not account:
            raise ValueError(f"Account '{account_name}' not found.")

        row_number = 0
        for line in content_lines:
            row_number += 1
            line = line.strip()
            if not line:
                continue

            raw_lines_buffer.append(line)

            if line == "^":
                if current_transaction:
                    # Deduplication using hash of raw lines
                    raw_content = "".join(raw_lines_buffer)
                    internal_id = hashlib.sha256(
                        raw_content.encode("utf-8")
                    ).hexdigest()

                    # Reset buffer for next transaction
                    raw_lines_buffer = []

                    try:
                        with transaction.atomic():
                            if Transaction.objects.filter(
                                internal_id=internal_id
                            ).exists():
                                self._increment_totals("skipped", 1)
                                self._log(
                                    "info",
                                    f"Skipped duplicate transaction from {filename}",
                                )
                                current_transaction = {}
                                continue

                            # Handle Account
                            if account:
                                current_transaction["account"] = account
                            else:
                                acc = Account.objects.filter(name=account_name).first()
                                if acc:
                                    current_transaction["account"] = acc
                                else:
                                    raise ValueError(
                                        f"Account '{account_name}' not found."
                                    )

                            current_transaction["internal_id"] = internal_id

                            # Handle Description/Memo mapping
                            if "memo" in current_transaction:
                                current_transaction["description"] = (
                                    current_transaction.pop("memo")
                                )

                            # Handle Payee mapping
                            entities = []
                            if "payee" in current_transaction:
                                payee_name = current_transaction.pop("payee")
                                # "Treat the payee (P) as the entity. Use existing or create"
                                entity, _ = TransactionEntity.objects.get_or_create(
                                    name=payee_name
                                )
                                entities.append(entity)

                            # Handle Label/Category
                            category = None
                            tags = []
                            if "label" in current_transaction:
                                label = current_transaction.pop("label")
                                if label.startswith("[") and label.endswith("]"):
                                    # Transfer: set label as description, ignore category/tags
                                    clean_label = label[1:-1]
                                    current_transaction["description"] = clean_label
                                else:
                                    parts = label.split(":")
                                    if parts:
                                        cat_name = parts[0].strip()
                                        if cat_name:
                                            category, _ = (
                                                TransactionCategory.objects.get_or_create(
                                                    name=cat_name
                                                )
                                            )

                                        if len(parts) > 1:
                                            for tag_name in parts[1:]:
                                                tag_name = tag_name.strip()
                                                if tag_name:
                                                    tag, _ = (
                                                        TransactionTag.objects.get_or_create(
                                                            name=tag_name
                                                        )
                                                    )
                                                    tags.append(tag)

                            current_transaction["category"] = category

                            # Create transaction
                            new_trans = Transaction.objects.create(
                                **current_transaction
                            )
                            if entities:
                                new_trans.entities.set(entities)
                            if tags:
                                new_trans.tags.set(tags)

                            self.import_run.transactions.add(new_trans)
                            self._increment_totals("successful", 1)

                    except Exception as e:
                        if not self.settings.skip_errors:
                            raise e
                        self._log(
                            "warning",
                            f"Error processing transaction in {filename}: {str(e)}",
                        )
                        self._increment_totals("failed", 1)

                    # Reset for next transaction
                    current_transaction = {}
                else:
                    # Empty transaction record (orphaned ^)
                    raw_lines_buffer = []
                    pass
                self._increment_totals("processed", 1)
                continue

            if line.startswith("!"):
                continue

            code = line[0]
            value = line[1:]

            if code == "D":
                try:
                    current_transaction["date"] = datetime.strptime(
                        value, self.settings.date_format
                    ).date()
                except ValueError:
                    self._log(
                        "warning",
                        f"Could not parse date '{value}' using format '{self.settings.date_format}' in {filename}",
                    )
                    if not self.settings.skip_errors:
                        raise ValueError(f"Invalid date format '{value}'")

            elif code == "T":
                try:
                    cleaned_value = value.replace(",", "")
                    amount = Decimal(cleaned_value)
                    if amount < 0:
                        current_transaction["type"] = Transaction.Type.EXPENSE
                        current_transaction["amount"] = abs(amount)
                    else:
                        current_transaction["type"] = Transaction.Type.INCOME
                        current_transaction["amount"] = amount
                except InvalidOperation:
                    self._log(
                        "warning", f"Could not parse amount '{value}' in {filename}"
                    )
                    if not self.settings.skip_errors:
                        raise ValueError(f"Invalid amount format '{value}'")

            elif code == "P":
                current_transaction["payee"] = value
            elif code == "M":
                current_transaction["memo"] = value
            elif code == "L":
                current_transaction["label"] = value
            elif code == "N":
                pass

    def _process_qif(self, file_path):
        def process_logic():
            if zipfile.is_zipfile(file_path):
                try:
                    with zipfile.ZipFile(file_path, "r") as zf:
                        for filename in zf.namelist():
                            if filename.lower().endswith(
                                ".qif"
                            ) and not filename.startswith("__MACOSX"):
                                self._log(
                                    "info", f"Processing QIF from ZIP: {filename}"
                                )
                                with zf.open(filename) as f:
                                    content = f.read().decode(self.settings.encoding)
                                    self._parse_and_import_qif(
                                        content.splitlines(), filename
                                    )
                except Exception as e:
                    raise ValueError(f"Error processing ZIP file: {str(e)}")
            else:
                with open(file_path, "r", encoding=self.settings.encoding) as f:
                    self._parse_and_import_qif(
                        f.readlines(), os.path.basename(file_path)
                    )

        if not self.settings.skip_errors:
            with transaction.atomic():
                process_logic()
        else:
            process_logic()

    def _validate_file_path(self, file_path: str) -> str:
        """
        Validates that the file path is within the allowed temporary directory.
        Returns the absolute path.
        """
        abs_path = os.path.abspath(file_path)
        if not abs_path.startswith(self.TEMP_DIR):
            raise ValueError(f"Invalid file path. File must be in {self.TEMP_DIR}")
        return abs_path

    def process_file(self, file_path: str):
        with cachalot_disabled():
            # Validate and get absolute path
            file_path = self._validate_file_path(file_path)

            self._update_status("PROCESSING")
            self.import_run.started_at = timezone.now()
            self.import_run.save(update_fields=["started_at"])

            self._log("info", "Starting import process")

            try:
                if isinstance(self.settings, version_1.CSVImportSettings):
                    self._process_csv(file_path)
                elif isinstance(self.settings, version_1.ExcelImportSettings):
                    self._process_excel(file_path)
                elif isinstance(self.settings, version_1.QIFImportSettings):
                    self._process_qif(file_path)

                self._update_status("FINISHED")
                self._log(
                    "info",
                    f"Import completed successfully. "
                    f"Successful: {self.import_run.successful_rows}, "
                    f"Failed: {self.import_run.failed_rows}, "
                    f"Skipped: {self.import_run.skipped_rows}",
                )

            except Exception as e:
                self._update_status("FAILED")
                self._log("error", f"Import failed: {str(e)}")
                raise Exception("Import failed")

            finally:
                self._log("info", "Cleaning up temporary files")
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                        self._log("info", f"Deleted temporary file: {file_path}")
                except OSError as e:
                    self._log("warning", f"Failed to delete temporary file: {str(e)}")

                self.import_run.finished_at = timezone.now()
                self.import_run.save(update_fields=["finished_at"])
